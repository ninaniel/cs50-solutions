import pytest
import project as p
from openpyxl import load_workbook


books =  [
        [1, "book1", "author1", "10.00$" ],
        [2, "book2", "author2", "20.00$"],
        [3, "book3", "author3", "30.00$" ],
        [4, "book4", "author4", "40.00$"],
    ]


def test_input_validation(monkeypatch):
    monkeypatch.setattr('builtins.input', lambda _: '1')
    assert p.validate_input() == 1

    monkeypatch.setattr('builtins.input', lambda _: '5')
    assert p.validate_input() == 5

    monkeypatch.setattr('builtins.input', lambda _: '7')
    with pytest.raises(SystemExit):
        p.validate_input()


def test_load_books():
    books_list = p.load_books("")
    assert isinstance(books_list, list)
    assert len(books_list) == 100


def test_buy_book(monkeypatch):
    monkeypatch.setattr('builtins.input', lambda _: '1')
    p.select_book(books)
    assert p.checkout_items == ["book1 - $10.0"]
    assert p.total_price == 10.0

    monkeypatch.setattr('builtins.input', lambda _: '2')
    p.select_book(books)
    assert p.checkout_items == ["book1 - $10.0","book2 - $20.0"]
    assert p.total_price == 30.0


def test_remove_item(monkeypatch):
    monkeypatch.setattr('builtins.input', lambda _: '1')
    p.remove_item()
    assert p.checkout_items == ['book2 - $20.0']
    assert p.total_price == 20.0

    p.remove_item()
    assert p.checkout_items == []
    assert p.total_price == 0.0


def test_next():
    p.next_page()
    assert p.start_row == 101
    assert p.end_row == 201
    assert p.page == 2

    p.next_page()
    assert p.start_row == 201
    assert p.end_row == 301
    assert p.page == 3

    p.next_page()
    assert p.start_row == 201
    assert p.end_row == 301
    assert p.page == 3

def test_previous():
    p.previous_page()
    assert p.start_row == 101
    assert p.end_row == 201
    assert p.page == 2

    p.previous_page()
    assert p.start_row == 1
    assert p.end_row == 101
    assert p.page == 1

    p.previous_page()
    assert p.start_row == 1
    assert p.end_row == 101
    assert p.page == 1


def test_empty_bag_options(monkeypatch):
    monkeypatch.setattr('builtins.input', lambda _: "no")
    with pytest.raises(SystemExit):
        p.empty_bag_options("testiing")


def test_cleaning_up_before_exiting():
    workbook = load_workbook("collections.xlsx")

    assert len(workbook.sheetnames) == 1

    with pytest.raises(SystemExit):
        p.exit("exiting")