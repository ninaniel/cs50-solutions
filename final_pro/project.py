import sys
import requests
from bs4 import BeautifulSoup
import random
from openpyxl import load_workbook
from tabulate import tabulate
from fpdf import FPDF
from datetime import datetime

start_row, end_row, page = 1, 101, 1 
checkout_items = []
total_price = 0
chosen_collections = []

def main():
    global chosen_collections
    greeting()

    while True:
        main_categories_menu()
        user_choice = validate_input()
           
        if not user_choice in chosen_collections:
            create_collection(user_choice)
            chosen_collections.append(user_choice)    

        while True:
            books = load_books(user_choice)
            shopping_menu()

            while True:
                choice_shopping = input("➡ Enter your choice: ").lower()
                choices = ["a", "b", "c", "d", "e"]
                if not choice_shopping in choices:
                    print("Invalid choice. Please enter a valid option.") 
                else:
                    break

            match choice_shopping:
                case "a":  # shopping process
                    select_book(books)

                    while True:
                        checkout_menu()
                        checkout_choice = input("➡ Enter your choice: ")

                        match checkout_choice:
                            # proceed to checkout
                            case '1': 
                                display_shopping_bag()
                                checkout()
                                break
                            # add item - reload last page 
                            case '2':         
                                break 
                            # remove item from the bag     
                            case '3':
                                display_shopping_bag()
                                remove_item()
                                if len(checkout_items) == 0:  
                                    empty_bag_options()
                                    break
                            # review shopping bag    
                            case '4':          
                                display_shopping_bag()
                                input("Press any key to go Back. ")
                            # cancel - emptying bag    
                            case '5':
                                empty_bag_options()
                                break                         
                            case _:
                                print("Invalid choice. Please enter a valid option.")

                case  "b":
                    next_page()
                case  "c":
                    previous_page()
                case  "d":  # return to main categories
                    break
                case  "e":
                    exit("Exiting the program. Thank you!")             


def validate_input():
    while True:
        try:
            user_input = int(input("➡ Your choice: "))

            if not 0 < user_input < 8:
                raise ValueError
            elif user_input == 6:
                while True:
                    try:
                        user_input = int(input("Please, enter the century (17/18/19/20/21): "))
                        if 17 <= user_input <= 21:
                            break
                    except ValueError:
                        pass            
            elif user_input == 7:
                exit("Exiting...\nHope You'll return soon.🤓")

            return user_input
        
        except ValueError:
                print("Invalid choice. Please enter a valid option.") 
    
def create_collection(collection):
    records = [["N", "📝 Title", "🤓 Author", "💲Price"]]
    topics = {
        1: "952.1001_Books_You_Must_Read_Before_You_Die",
        2: "2700.Science_Fiction_and_Fantasy_Must_Reads",
        3: "1362.Best_History_Books_",
        4: "281.Best_Memoir_Biography_Autobiography",
        5: "7616.Motivational_and_Self_Improvement_Books",
        17:"53", 18:"30", 19:"16", 20:"6", 21: "7"
    }

    print("⏳ Loading 100 books from total of 300\n⏳")

    for n in range(3):
        url = f"https://www.goodreads.com/list/show/{topics[collection]}?page={n+1}"
        r = requests.get(url)
        c = r.text

        soup = BeautifulSoup(c, 'html.parser')
        data = soup.find_all("tr")

        for index, row in enumerate(data, start=1):
            book_info = row.find_all("span")
            title = book_info[0].text
            author = book_info[3].text
            price = f"{round(random.uniform(9.0, 42.0), 2)}$"
            item = [index, title, author, price]
            records.append(item)

    workbook = load_workbook("collections.xlsx")
    sheet_title = f"Sheet{collection}"
    workbook.create_sheet(title=sheet_title)     
    sheet = workbook[sheet_title]
    
    for item in records:
        sheet.append(item) 

    workbook.save("collections.xlsx")
    workbook.close

def load_books(n):
    workbook = load_workbook("collections.xlsx")
    sheet = workbook[f"Sheet{n}"]

    head_row = list(sheet.values)[0]
    shop_list = list(sheet.values)[start_row:end_row]
    widths = [None, 100, None, None]

    print(tabulate(shop_list, headers=head_row, tablefmt="grid", maxcolwidths=widths),
          f"\n📌 Page - {page} / 3")

    return shop_list
      
def select_book(books):
    global checkout_items, total_price
    while True:
        try:
            book_choice = int(input("Enter the number of the book you want to buy: "))
            if not book_choice in range(1,101):
                raise ValueError 
            else:
                break
        except ValueError:
            print("⚠ Please, enter the valid number of the book.")
    
    book_index = book_choice - 1
    book_title = books[book_index][1]
    book_price = float(books[book_index][3].removesuffix("$"))

    chosen_book = f"{book_title} - ${book_price}"
    checkout_items.append(chosen_book)
    total_price = round((total_price + book_price), 2)

    print(f"\nYou have added: 📖 {chosen_book}")

def checkout():
    while True:
        pay_choice = input(f"📌 Total Price: ${total_price}\n\nConfirm payment? (yes/no): ").lower()                           
        match pay_choice:
            case 'yes':
                create_invoice()
                exit("Invoice created ✔. Thank you for shopping!")
            case "no":
                empty_bag_options(prompt="Payment canceled. Shopping bag is empty.")
                return 
            case _:
                continue

def display_shopping_bag():
    print("📌 Items:\n", tabulate(enumerate(checkout_items, start=1)))

def remove_item():
    global checkout_items, total_price
    while True:
        try:
            index = int(input("Choose book to remove: "))
            book = checkout_items[index-1]
            book_price = float(book.split(" - $")[1])
                                            
            print(f"\n ✔'{book}' Removed.")
            checkout_items.remove(book)

            total_price = round((total_price - book_price), 2)
            return
            
        except (ValueError, IndexError):
            print("Invalid choice. Please enter a valid option.")

def next_page():
    global start_row, end_row, page
    start_row += 100
    end_row += 100
    page += 1
    if page > 3:   
        start_row, end_row, page = 201, 301, 3  # reloading last page
    
def previous_page():
    global start_row, end_row, page
    start_row -= 100
    end_row -= 100
    page -= 1
    if page == 0:                
        start_row, end_row, page = 1, 101, 1  # reloading first page 
    

def greeting():
    print("📚 Welcome to our Book Shop 📚",
          "\n🧐 Searching for some good books to enjoy yourself?",
            "⬇ Have a look at our collections! ⬇", sep="\n")

def main_categories_menu():
    print(  
            "\n➡ Choose a category:"
            "\n1. Books You Must Read Before You Die",
            "2. Science Fiction & Fantasy Books",
            "3. History Books",
            "4. Memoir / Biography / Autobiography",
            "5. Motivational and Self-Improvement Books",
            "6. Best Books by Century",
            "7. ❌ Exit", sep = "\n") 
    
def shopping_menu():
    print("\n⚠ Actions:\n",
          "a) 📕 Choose a book to buy",
          "b) 📜 Next page",
          "c) 📜 Previous page",
          "d) 📚 Go Back to Main Collections",
          "e) ❌ Exit", sep= "\n")
    
def checkout_menu():
    print(f"\n⚠ Items in the shopping bag: {len(checkout_items)}",
          f"⚠ Total Price: ${total_price}",
          "\n1. ✅ Check-out",
          "2. ➕ Add item",
          "3. ➖ Remove item",
          "4. 🛒 Display shopping bag",
          "5. ❌ Cancel", sep= "\n")
    
def empty_bag_options(prompt="Shopping bag is empty."):
    global total_price, checkout_items
    total_price = 0
    checkout_items = []

    while True:
        choice = input(f"{prompt}\n🛒 Continue shopping? (yes/no): ").lower()
        
        if choice == 'yes':           
            return
        elif choice == 'no':
            exit("Exiting...")
        else:
            continue

def exit(message):
    workbook = load_workbook("collections.xlsx")
    for sheet in workbook.sheetnames:
        if sheet != "Sheet":  # keep default sheet
            workbook.remove(workbook[sheet])
    workbook.save("collections.xlsx")

    sys.exit(message)

def create_invoice():
    while True:
        name = input("⚠ Please, enter your personal information for Invoice\nYour name: ").title()
        address = input("Your address: ").capitalize()
        if name and address:
            break
           
    pdf = FPDF()
    pdf.add_page()

    # title
    pdf.set_font("Arial", "B", 24)
    pdf.cell(0, 20, "INVOICE", ln=True, align="C")

    # store name
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, "Book Store", ln=True, align="C")

    # invoice, order info 
    pdf.ln(10)
    pdf.set_font("Arial", "", 12)
    pdf.cell(0, 10, f"Invoice #: {str(random.randint(100000, 999999))}", ln=False, align="L")
    pdf.cell(0, 10, f"Order #: {str(random.randint(100000, 999999))}", ln=True, align="R")
    pdf.cell(0, 10, f"Invoice Date: {datetime.now().strftime('%Y-%m-%d')}", ln=False, align="L")
    pdf.cell(0, 10, f"Order Date: {datetime.now().strftime('%Y-%m-%d')}", ln=True, align="R")

    # costumer info
    pdf.ln(10)
    pdf.cell(0, 7, "Invoice To", ln=True)
    pdf.cell(0, 7, f"Customer Name: {name}", ln=True)
    pdf.cell(0, 7, f"Address: {address}", ln=True)

    # product table
    pdf.ln(10)
    pdf.set_font("Arial", "B", 12)
    pdf.set_fill_color(0, 128, 0)
    page_width = pdf.w - (pdf.l_margin + pdf.r_margin)
    pdf.cell(0.75*page_width, 10, "Product Description", 1, 0, "C", fill=True)
    pdf.cell(0.25*page_width, 10, "Unit Price", 1, 1, "C", fill=True)

    # product details
    pdf.set_font("Arial", "", 12)
    for item in checkout_items:
        pdf.cell(0.75*page_width,10, item.split(" - ")[0], 1, 0, "L")
        pdf.cell(0.25*page_width,10, item.split(" - ")[1], 1, 1, "R")

    # total price
    pdf.ln(10)
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, f"Total Price: {total_price}", ln=True, align="R")

    pdf.output("invoice.pdf")   


if __name__ == "__main__":
    main()